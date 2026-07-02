"""Shared export toolkit: PDF (reportlab), Excel (openpyxl), CSV, JSON.

Both modules describe a report as a list of blocks:
  {"type": "title",    "text": str}
  {"type": "heading",  "text": str}
  {"type": "paragraph","text": str}
  {"type": "kv",       "items": [[label, value], ...]}
  {"type": "table",    "columns": [...], "rows": [[...], ...]}
and this module renders that description into any requested format.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .config import EXPORT_DIR, ensure_dirs

ACCENT = colors.HexColor("#6d5df6")
DARK = colors.HexColor("#17182b")
GREY = colors.HexColor("#8a8fa8")


def _stamp(prefix: str, ext: str) -> Path:
    ensure_dirs()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return EXPORT_DIR / f"{prefix}-{ts}.{ext}"


def _fmt(v: Any) -> str:
    if isinstance(v, float):
        return f"{v:,.4f}".rstrip("0").rstrip(".") if abs(v) < 1e6 else f"{v:,.0f}"
    return str(v)


# ---------------------------------------------------------------------------- PDF
def export_pdf(prefix: str, blocks: list[dict[str, Any]]) -> Path:
    path = _stamp(prefix, "pdf")
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=prefix,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], textColor=DARK, spaceAfter=6)
    head_style = ParagraphStyle("H", parent=styles["Heading2"], textColor=ACCENT, spaceBefore=10)
    body_style = ParagraphStyle("B", parent=styles["BodyText"], textColor=DARK, leading=14)
    flow: list[Any] = []
    for b in blocks:
        t = b["type"]
        if t == "title":
            flow.append(Paragraph(b["text"], title_style))
            flow.append(Paragraph(
                datetime.now(timezone.utc).strftime("Generated %Y-%m-%d %H:%M UTC"),
                ParagraphStyle("sub", parent=body_style, textColor=GREY)))
            flow.append(Spacer(1, 6))
        elif t == "heading":
            flow.append(Paragraph(b["text"], head_style))
        elif t == "paragraph":
            flow.append(Paragraph(b["text"], body_style))
        elif t == "kv":
            rows = [[str(k), _fmt(v)] for k, v in b["items"]]
            tbl = Table(rows, colWidths=[70 * mm, 90 * mm])
            tbl.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("TEXTCOLOR", (0, 0), (0, -1), GREY),
                ("TEXTCOLOR", (1, 0), (1, -1), DARK),
                ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#e3e4ee")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            flow.append(tbl)
        elif t == "table":
            data = [[str(c) for c in b["columns"]]] + [[_fmt(c) for c in r] for r in b["rows"]]
            tbl = Table(data, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d5d7e4")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4fb")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            flow.append(tbl)
            flow.append(Spacer(1, 6))
    doc.build(flow)
    return path


# -------------------------------------------------------------------------- Excel
def export_xlsx(prefix: str, sheets: dict[str, list[dict[str, Any]]]) -> Path:
    """sheets: {sheet_name: blocks} — kv and table blocks are rendered."""
    path = _stamp(prefix, "xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="6D5DF6")
    for name, blocks in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        row = 1
        for b in blocks:
            t = b["type"]
            if t in ("title", "heading"):
                c = ws.cell(row=row, column=1, value=b["text"])
                c.font = Font(bold=True, size=14 if t == "title" else 11, color="17182B")
                row += 2
            elif t == "paragraph":
                ws.cell(row=row, column=1, value=b["text"])
                row += 2
            elif t == "kv":
                for k, v in b["items"]:
                    ws.cell(row=row, column=1, value=str(k)).font = Font(color="8A8FA8")
                    ws.cell(row=row, column=2, value=v if isinstance(v, (int, float)) else str(v))
                    row += 1
                row += 1
            elif t == "table":
                for j, col in enumerate(b["columns"], start=1):
                    c = ws.cell(row=row, column=j, value=str(col))
                    c.font = head_font
                    c.fill = head_fill
                    c.alignment = Alignment(horizontal="center")
                row += 1
                for r in b["rows"]:
                    for j, v in enumerate(r, start=1):
                        ws.cell(row=row, column=j, value=v if isinstance(v, (int, float)) else str(v))
                    row += 1
                row += 1
        for j in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(j)].width = 18
    wb.save(path)
    return path


# ----------------------------------------------------------------------- CSV/JSON
def export_csv(prefix: str, columns: list[str], rows: list[list[Any]]) -> Path:
    path = _stamp(prefix, "csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(columns)
        w.writerows(rows)
    return path


def export_json(prefix: str, payload: Any) -> Path:
    path = _stamp(prefix, "json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=lambda o: o.item() if hasattr(o, "item") else str(o))
    return path
